import traceback
import yaml
import requests
import openpyxl
import jpholiday
import output_jira_chart
from requests.auth import HTTPBasicAuth
from datetime import datetime, timedelta

# yamlファイルの読み込み
with open('config.yml', 'r', encoding='utf-8') as yml:
    config = yaml.safe_load(yml)

# 定数の宣言
URL = config['url']
AUTH = HTTPBasicAuth(config['username'], config['password'])
PARENT_JQL = config['sprint_id'] + config['parent_jql']
CHILD_JQL = config['sprint_id'] + config['child_jql']
WORK_START_TIME = datetime.strptime(config['work_start_time'], '%H:%M').time()
WORK_END_TIME = datetime.strptime(config['work_end_time'], '%H:%M').time()
BREAK_START_TIME = datetime.strptime(config['break_start_time'], '%H:%M').time()
BREAK_END_TIME = datetime.strptime(config['break_end_time'], '%H:%M').time()
FILE_PATH = config['file_path']

class Issue:
    def __init__(self, issue):
        self.data = issue
        self.fields = issue.get('fields')
        
    @property
    def key(self):
        return self.data.get('key')

    @property
    def id(self):
        return self.data.get('id')
    
    @property
    def epic_key(self):
        return self.fields.get('customfield_10001', '')
        
    @property
    def sprint(self):
        sprints = [label for label in self.fields.get('labels', []) if 'Sprint' in label]
        if len(sprints) == 0:
            return ''
        else:
            sprints.sort(reverse=True)        
            return sprints[0]
        
    @property
    def issue_type(self):
        return self.fields.get('issuetype').get('name')
    
    @property
    def status(self):
        return self.fields.get('status').get('name')
    
    @property
    def summary(self):
        return self.fields.get('summary')
    
    @property
    def story_point(self):
        if self.fields.get('customfield_xxxxx') is not None:
            return self.fields.get('customfield_xxxxx')
        else:
            return 0
    
    @property
    def category(self):
        category = self.fields.get('customfield_xxxxx', None)
        if category is not None:
            return category.get('value', '未割当')
        else:
            return '未割当'
        
    @property
    def team(self):
        return 'ほげほげチーム'
    
    @property
    def assignee(self):
        assignee = self.fields.get('assignee', None)
        if assignee is not None:
            return assignee.get('displayName', '未割当')
        else:
            return '未割当'
        
    @property
    def assignees(self):
        if 'customfield_xxxxx' in self.fields and self.fields.get('customfield_xxxxx') is not None:
            assignees = []
            for assignee in self.fields.get('customfield_xxxxx'):
                assignees.append(assignee.get('displayName', ''))
            return ';'.join(assignees)
        else:
            return '未割当'
        
    @property
    def created_date(self):
        return self.time_format(self.fields.get('created'))

    @property
    def start_date(self):
        date = self.time_format(self.fields.get('customfield_xxxxx'))
        if date == '':
            return datetime.now()
        else:
            return date
    
    @property
    def end_date(self):
        date = self.time_format(self.fields.get('customfield_xxxxx'))
        if date == '':
            return datetime.now()
        else:
            return date
    
    @property
    def release(self):
        if self.category == 'リリース作業':
            return 1
        else:
            return 0
    
    def time_format(self, time):
        if time is None or time == '':
            return ''
        else:
            return datetime.strptime(time, '%Y-%m-%dT%H:%M:%S.%f%z')
    
    def parent_duct(self):
        return {
            'key':self.key,
            'id':self.id,
            'epic_key':self.epic_key,
            'sprint':self.sprint,
            'issue_type':self.issue_type,
            'status':self.status,
            'summary':self.summary,
            'story_point':self.story_point,
            'team':self.team,
            'assignee':self.assignee,
            'assignees':self.assignees,
            'created_date':self.created_date,
            'start_date':self.start_date,
            'end_date':self.end_date,
        }
    
    def child_duct(self):
        return {
            'key':self.key,
            'id':'',
            'sprint':self.sprint,
            'issue_type':self.issue_type,
            'status':self.status,
            'summary':self.summary,
            'category':self.category,
            'team':self.team,
            'assignee':self.assignee,
            'assignees':self.assignees,
            'start_date':self.start_date,
            'end_date':self.end_date,
        }
        
def get_issues(type, jql):
    respons = requests.get(URL + type, auth=AUTH, params={'jql':jql})
    if type == 'search':
        jira_data = respons.json()['issues']
    else:
        jira_data = respons.json()
    
    return jira_data

def get_filed_issues(parent_issues):
    parent_issues_data = []
    child_issues_data = []
    i = 0
    
    for i, parent_issue in enumerate(parent_issues):
        print(f'JIRAの課題から値を取得中... ({i + 1}/{len(parent_issues)}件)')
        
        # 初期化
        total_time_spent = 0
        
        # 親課題の情報を取得
        parent_issue_data = Issue(parent_issue).parent_duct()
        
        # 親課題のidを使って子課題を取得
        child_issues = get_issues('search', CHILD_JQL + parent_issue_data['id']) 
        for child_issue in child_issues:
            # 子課題の情報を取得
            child_issue_data = Issue(child_issue).child_duct()
            child_issue_data['id'] = parent_issue_data['id']
        
            # 消費時間の計算
            time_spent = work_time_spent(child_issue_data['start_date'], child_issue_data['end_date'])
            total_time_spent = total_time_spent + time_spent
            
            # JIRAで保持ってない子課題のデータを追加
            child_issue_data = {**child_issue_data, **{'process_time':time_spent}}
                
            child_issues_data.append(child_issue_data)

        # プロセスタイムを計算
        if total_time_spent != 0 and parent_issue_data['story_point'] != 0:
            process_time = float(round(total_time_spent / parent_issue_data['story_point'], 2))
        else:
            process_time = 0

        # サイクルタイムを計算
        cycle_time = float(round(work_time_spent(parent_issue_data['start_date'], parent_issue_data['end_date']), 2))
        if cycle_time != 0 and parent_issue_data['story_point'] != 0:
            cycle_time_story_point = float(round(cycle_time / parent_issue_data['story_point'], 2))
        elif cycle_time != 0 and parent_issue_data['story_point'] == 0:
            cycle_time_story_point = cycle_time
        else:
            cycle_time_story_point = 0
            
        # リードタイムを計算
        read_time = float(round(work_time_spent(parent_issue_data['created_date'], parent_issue_data['end_date']), 2))

        # JIRAで保持ってない親課題のデータを追加
        parent_issue_data_add = {
            'time_spent':round(total_time_spent, 2),
            'process_time':process_time,
            'cycle_time_story_point':cycle_time_story_point,
            'cycle_time':cycle_time,
            'read_time':read_time,
            'release_created_date':'',
            'release_start_dete':'',
            'release_end_date':'',
            'release_count':0,
        }
        parent_issue_data = {**parent_issue_data, **parent_issue_data_add}        
        
        parent_issues_data.append(parent_issue_data)
    
    return parent_issues_data, child_issues_data

def work_time_spent(start_date, end_date):
    if start_date == '' or end_date == '':
        return 0
    else:
        # オフセット有無でエラーとなるのを回避
        current_date = start_date.replace(tzinfo=None)
        end_date_obj = end_date.replace(tzinfo=None)
    
    total_work_hours = timedelta(0)
    
    # 平日10:00~12:00,13:00~19:00を作業時間と定義し、作業した時間を計算
    while current_date.date() <= end_date_obj.date():
        # 土日と祝日はスキップ
        if current_date.weekday() >= 5 or jpholiday.is_holiday(current_date.date()):
            current_date += timedelta(1)
            continue
        
        # 作業開始時間と作業終了時間を計算
        daily_work_start_time = max(current_date, datetime.combine(current_date.date(), WORK_START_TIME))
        daily_work_end_time = min(end_date_obj, datetime.combine(current_date.date(), WORK_END_TIME))
        
        # 作業時間から休憩時間を差し引く
        if daily_work_start_time < datetime.combine(current_date.date(), BREAK_END_TIME) and daily_work_end_time > datetime.combine(current_date.date(), BREAK_START_TIME):
            if daily_work_start_time < datetime.combine(current_date.date(), BREAK_START_TIME):
                total_work_hours += datetime.combine(current_date.date(), BREAK_START_TIME) - daily_work_start_time
            if daily_work_end_time > datetime.combine(current_date.date(), BREAK_END_TIME):
                total_work_hours += daily_work_end_time - datetime.combine(current_date.date(), BREAK_END_TIME)
        else:
            # 休憩時間がない場合
            if daily_work_start_time < daily_work_end_time:
                total_work_hours += daily_work_end_time - daily_work_start_time
        
        # 翌日の作業時間を計算
        current_date = datetime.combine(current_date.date() + timedelta(1), WORK_START_TIME)
    
    return total_work_hours.total_seconds() / 3600

def write_to_excel(parent_issues_data, child_issues_data):
    print('Excelに書き込み中...')
    
    wb = openpyxl.load_workbook(FILE_PATH)
    
    # 親課題のデータをExcelに書き込む
    ws_parent = wb['parent_data']
    for data in parent_issues_data:
        row = ws_parent.max_row + 1
        for row_check in range(2, ws_parent.max_row + 1):
            if ws_parent.cell(row=row_check, column=1).value == data['key']:
                row = row_check
                break
            
        for col, key in enumerate(data, start=1):
            value = data[key] 
            if col == 2:
                value = int(value)
            elif col >= 12 and col <= 14 and value != '':
                value = value.replace(tzinfo=None)
            
            ws_parent.cell(row=row, column=col, value=value)
    
    # 子課題のデータをExcelに書き込む
    ws_child = wb['child_data']
    for data in child_issues_data:
        row = ws_child.max_row + 1
        for row_check in range(2, ws_child.max_row + 1):
            if ws_child.cell(row=row_check, column=1).value == data['key']:
                row = row_check
                break
            
        for col, key in enumerate(data, start=1):
            value = data[key] 
            if col == 2:
                value = int(value)
            elif col >= 11 and col <= 12 and value != '':
                value = value.replace(tzinfo=None)
            
            ws_child.cell(row=row, column=col, value=value)

    wb.save(FILE_PATH)

def main():
    try:
        # 課題のデータを取得
        issues = get_issues('search', PARENT_JQL)
        parent_issues_data, child_issues_deta = get_filed_issues(issues)
        
        # Excelに書き込み
        write_to_excel(parent_issues_data, child_issues_deta)
        
        # グラフを出力
        output_jira_chart.output_chart(parent_issues_data, child_issues_deta)
                
        print('success:親課題 ' + str(len(parent_issues_data)) + '件/子課題 ' + str(len(child_issues_deta)) + '件')
    
    except Exception as e:
        print(traceback.format_exc())
    
if __name__ == '__main__':
    main()